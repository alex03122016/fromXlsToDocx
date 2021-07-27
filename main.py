import docx
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column, Integer, String, Date
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

def fromXlxsToDocx():
    #https://leportella.com/sqlalchemy-tutorial/
    #create tables models for sqlite3 database
    Base = declarative_base()

    class Schueler(Base):
        __tablename__ = 'Schueler'

        id = Column(Integer, primary_key=True)
        name = Column(String)
        vorname = Column(String)
        geburtsdatum = Column(Date)

    class Schuljahr(Base):
        __tablename__ = 'Schuljahr'
        id = Column(Integer, primary_key=True)
        Schuljahrbeginn = Column(Integer)
        Schuljahrende = Column(Integer)
        Klasse = Column(String)

    class Fehlzeiten(Base):
        __tablename__= 'Fehlzeiten'
        id = Column(Integer, primary_key=True)
        T_ins = Column(Integer)
        T_ue = Column(Integer)
        Stu = Column(Integer)
        Stu_ue = Column(Integer)
        v = Column(Integer)

    class Noten(Base):
        __tablename__ = 'Noten'
        id = Column(Integer, primary_key=True)
        Deu_ins = Column(Integer)
        Deu_m = Column(Integer)
        Deu_s = Column(Integer)
        Englisch = Column(Integer)
        GeWi_ins = Column(Integer)
        Geo = Column(Integer)
        Geschichte = Column(Integer)
        politischeBildung = Column(Integer)
        Ethik = Column(Integer)
        WAT = Column(Integer)
        Ma = Column(Integer)
        NaWi_ins = Column(Integer)
        Bio = Column(Integer)
        Che = Column(Integer)
        Phy = Column(Integer)
        Ku = Column(Integer)
        Mu = Column(Integer)
        Sp = Column(Integer)

    class Convert(Base):
        __tablename__ = 'convert'
        id = Column(Integer, primary_key=True)
        field = Column(String)
        row_Zeugnisliste = Column(Integer)
        column_Zeugnisliste = Column(Integer)
        table_zeugnisliste = Column(Integer)
        row_calctable = Column(Integer)
        column_calctable = Column(Integer)

    #open sqlite
    engine = create_engine('sqlite:///test.db')
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()

    #open .xlxs
    wb = Workbook('test.xlxs')
    wb = openpyxl.load_workbook('test.xlsx')
    ws = wb.active
    sheet = wb['Zeugnisnoten_Abschluss']

    #open .docx
    savepath= "Zegnislisteneu.docx"
    doc = docx.Document('Zeugnisliste.docx')

    #transfer of data
    for i in range(1,221):
        #get coordinates data from Convert sqlite3
        data = session.query(Convert).filter_by(id=i).all()
        data = data[0]
        #get values  of fields in .xlxs file
        sourceValue = sheet.cell(row=data.row_calctable+1, column=data.column_calctable).value
        #get coordinates of field in .docx file
        r = data.row_Zeugnisliste
        c = data.column_Zeugnisliste
        tableDocx = doc.tables[data.table_zeugnisliste]
        targetCell = tableDocx.cell(r, c)
        targetCellValue = targetCell.text
        #write data  to .docxf file
        targetCell.text = str(sourceValue)

    doc.save(savepath)

if __name__ == "__main__":
    fromXlxsToDocx()
